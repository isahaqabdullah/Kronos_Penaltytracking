import os
import logging
from datetime import datetime, timezone

# Directory to store exported session JSONs (if needed)
# Use absolute path to avoid issues in Docker
SESSION_EXPORT_DIR = os.path.abspath(os.environ.get("SESSION_EXPORT_DIR", "session_exports"))
os.makedirs(SESSION_EXPORT_DIR, exist_ok=True)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# --- Utility Functions ---

def safe_filename(name: str) -> str:
    """
    Sanitize a string to be filesystem-safe.
    Replaces all non-alphanumeric characters with underscores.
    """
    return "".join(c if c.isalnum() or c in ("_", "-") else "_" for c in name)

def timestamp_now(fmt: str = "%Y%m%d_%H%M%S") -> str:
    """
    Return a UTC timestamp string for filenames or logs.
    Default format: YYYYMMDD_HHMMSS
    """
    return datetime.utcnow().strftime(fmt)

def utc_now() -> datetime:
    """
    Return current UTC datetime with timezone info.
    """
    return datetime.now(timezone.utc)

def export_session_data(session_name: str, data: dict) -> str:
    """
    Export session data to a JSON file in SESSION_EXPORT_DIR.
    Returns the path of the saved file.
    """
    import json
    filename = f"{safe_filename(session_name)}_{timestamp_now()}.json"
    path = os.path.join(SESSION_EXPORT_DIR, filename)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)
    logger.info(f"Session data exported to {path}")
    return path

def load_session_data(file_path: str) -> dict:
    """
    Load session data from a JSON file.
    """
    import json
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Session file '{file_path}' does not exist.")
    with open(file_path, "r", encoding="utf-8") as f:
        return json.load(f)

def export_session_csv(session_name: str, infringements: list, session_info: dict = None) -> str:
    """
    Export session data to CSV format.
    Returns the path of the saved file.
    """
    import csv
    filename = f"{safe_filename(session_name)}_{timestamp_now()}.csv"
    path = os.path.join(SESSION_EXPORT_DIR, filename)
    
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        
        # Write session info header
        if session_info:
            writer.writerow(["Session Information"])
            writer.writerow(["Name", session_info.get("name", "")])
            writer.writerow(["Status", session_info.get("status", "")])
            writer.writerow(["Started At", session_info.get("started_at", "")])
            writer.writerow([])  # Empty row
        
        # Write infringements header
        writer.writerow(["Infringements"])
        writer.writerow([
            "ID", "Kart Number", "Turn Number", "Description", "Observer",
            "Warning Count", "Penalty Due", "Penalty Description", "Penalty Taken", "Timestamp"
        ])
        
        # Write infringement data
        for inf in infringements:
            writer.writerow([
                inf.get("id", ""),
                inf.get("kart_number", ""),
                inf.get("turn_number", ""),
                inf.get("description", ""),
                inf.get("observer", ""),
                inf.get("warning_count", ""),
                inf.get("penalty_due", ""),
                inf.get("penalty_description", ""),
                inf.get("penalty_taken", ""),
                inf.get("timestamp", "")
            ])
        
        # Write history if available
        has_history = any(inf.get("history") for inf in infringements)
        if has_history:
            writer.writerow([])  # Empty row
            writer.writerow(["Infringement History"])
            writer.writerow([
                "Infringement ID", "Action", "Performed By", "Observer", "Details", "Timestamp"
            ])
            
            for inf in infringements:
                for hist in inf.get("history", []):
                    writer.writerow([
                        inf.get("id", ""),
                        hist.get("action", ""),
                        hist.get("performed_by", ""),
                        hist.get("observer", ""),
                        hist.get("details", ""),
                        hist.get("timestamp", "")
                    ])
    
    logger.info(f"Session data exported to CSV: {path}")
    return path

def export_session_excel(session_name: str, infringements: list, session_info: dict = None) -> str:
    """
    Export session data to Excel format (.xlsx).
    Returns the path of the saved file.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    filename = f"{safe_filename(session_name)}_{timestamp_now()}.xlsx"
    path = os.path.join(SESSION_EXPORT_DIR, filename)
    
    wb = Workbook()
    
    # === Sheet 1: Infringements ===
    ws = wb.active
    ws.title = "Infringements"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Write session info
    if session_info:
        ws["A1"] = "Session Information"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = "Name:"
        ws["B2"] = session_info.get("name", "")
        ws["A3"] = "Status:"
        ws["B3"] = session_info.get("status", "")
        ws["A4"] = "Started At:"
        ws["B4"] = session_info.get("started_at", "")
        ws.append([])  # Empty row
    
    # Write infringements header
    start_row = 6 if session_info else 1
    headers = [
        "ID", "Kart Number", "Turn Number", "Description", "Observer",
        "Warning Count", "Penalty Due", "Penalty Description", "Penalty Taken", "Timestamp"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write infringement data
    for inf in infringements:
        row = [
            inf.get("id", ""),
            inf.get("kart_number", ""),
            inf.get("turn_number", ""),
            inf.get("description", ""),
            inf.get("observer", ""),
            inf.get("warning_count", ""),
            inf.get("penalty_due", ""),
            inf.get("penalty_description", ""),
            inf.get("penalty_taken", ""),
            inf.get("timestamp", "")
        ]
        ws.append(row)
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # === Sheet 2: History ===
    has_history = any(inf.get("history") for inf in infringements)
    if has_history:
        ws2 = wb.create_sheet("History")
        history_headers = [
            "Infringement ID", "Action", "Performed By", "Observer", "Details", "Timestamp"
        ]
        
        for col, header in enumerate(history_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for inf in infringements:
            for hist in inf.get("history", []):
                row = [
                    inf.get("id", ""),
                    hist.get("action", ""),
                    hist.get("performed_by", ""),
                    hist.get("observer", ""),
                    hist.get("details", ""),
                    hist.get("timestamp", "")
                ]
                ws2.append(row)
        
        # Auto-adjust column widths for history sheet
        for col in range(1, len(history_headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in ws2[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws2.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(path)
    logger.info(f"Session data exported to Excel: {path}")
    return path
